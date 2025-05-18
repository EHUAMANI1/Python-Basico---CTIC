#Identificar Variables

import pandas as pd
import os 
from openpyxl import load_workbook

# Ruta al archivo de origen
base_dir = os.getcwd()

archivo_consolidado = os.path.join(base_dir, "Data", "Entrada", "Albamar Bea Sabana ABR25.xlsm")
print("Cargando archivo:", archivo_consolidado)

def mostrar_nombres_variables(ruta_archivo, hoja_objetivo):
    # Cargar el archivo
    wb = load_workbook(ruta_archivo, data_only=True, keep_links=False)

    # Verificar que la hoja existe
    if hoja_objetivo not in wb.sheetnames:
        print(f"❌ La hoja '{hoja_objetivo}' no existe en el archivo.")
        return

    # Obtener la hoja específica
    ws = wb[hoja_objetivo]
    
    print(f"\n🔍 Buscando nombres de variables en la fila 5 de la hoja: '{hoja_objetivo}'\n")
    
    # Bandera para verificar si se encuentran variables
    variables_en_hoja = False
    
    # Iterar sobre la fila 5, comenzando desde la columna B (columna 2)
    for celda in ws[5][1:]:  # Se recorre la fila 5, desde la columna B (índice 1)
        # Si la celda tiene un valor (es decir, es una variable)
        if celda.value:
            print(f"📌 Nombre de la variable: {celda.value}")
            variables_en_hoja = True

    if not variables_en_hoja:
        print(f"⚠️ No se encontraron variables en la fila 5 de la hoja '{hoja_objetivo}'.")

# Llamar a la función para mostrar todos los nombres de variables en la fila 5 de la hoja 'Lista Precios'
mostrar_nombres_variables(archivo_consolidado, "Lista Precios")