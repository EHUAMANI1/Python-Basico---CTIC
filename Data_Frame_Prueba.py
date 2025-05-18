#Data_Frame_Prueba

import os
import pandas as pd
from openpyxl import load_workbook

# Ruta base
base_dir = os.getcwd()

# Archivo de entrada y salida
archivo_consolidado = os.path.join(base_dir, "Data", "Entrada", "Albamar Bea Sabana ABR25.xlsm")
archivo_salida = os.path.join(base_dir, "Data", "Salida", "FRAME.xlsx")

# Variables deseadas
variables_deseadas = [
    "Cod.", "No.", "Dorm.", "Modelo", "Tipo", "Vista", "Piso",
    "Area c/T (m2)", "Area s/T (m2)", "Status", "Cliente",
    "Precio", "Precio Proyectado"
]

# Cargar archivo Excel y hoja
wb = load_workbook(archivo_consolidado, data_only=True)
ws = wb["Lista Precios"]

# Leer encabezados desde fila 5, columna B (índice 1)
headers = [cell.value for cell in ws[5][1:] if cell.value in variables_deseadas]
column_indices = [i for i, cell in enumerate(ws[5][1:], start=1) if cell.value in variables_deseadas]

# Leer datos desde fila 6
data = []
for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
    fila = []
    for i in column_indices:
        valor = row[i].value
        fila.append(valor if valor not in ["", None] else None)
    if any(fila):  # Solo agregar si al menos una celda tiene valor
        data.append(fila)

# Crear DataFrame
df = pd.DataFrame(data, columns=headers)

# Exportar a Excel
df.to_excel(archivo_salida, index=False)
print(f"✅ Archivo exportado correctamente en:\n{archivo_salida}")
